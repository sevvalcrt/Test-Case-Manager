import os
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database.database import Database


# =====================================================
# TÜRKÇE KARŞILIKLAR
# =====================================================

STATUS_TRANSLATIONS = {

    "Yeni": "Yeni",

    "Beklemede": "Beklemede",

    "Devam Ediyor": "Devam Ediyor",

    "Başarılı": "Başarılı",

    "Başarısız": "Başarısız",

    "Engellendi": "Engellendi",

    # Eski İngilizce değerler

    "Not Run": "Çalıştırılmadı",

    "Pass": "Başarılı",

    "Fail": "Başarısız",

    "Blocked": "Engellendi"
}


MODULE_TRANSLATIONS = {

    "Login": "Giriş",

    "Search": "Arama",

    "Cart": "Sepet",

    "Product": "Ürün",

    "Checkout": "Ödeme",

    "Profile": "Profil"
}


# =====================================================
# PROJE YOLLARI
# =====================================================

PROJECT_ROOT = os.path.dirname(
    os.path.dirname(
        os.path.abspath(__file__)
    )
)


TEMPLATE_PATH = os.path.join(
    PROJECT_ROOT,
    "excel",
    "template.xlsx"
)


EXPORT_PATH = os.path.join(
    PROJECT_ROOT,
    "export.xlsx"
)


class ExcelExporter:

    # =================================================
    # INIT
    # =================================================

    def __init__(
        self,
        db=None
    ):

        if db is not None:

            self.db = db
            self.owns_db = False

        else:

            self.db = Database()
            self.owns_db = True

    # =================================================
    # STATUS STYLE
    # =================================================

    def apply_status_style(
        self,
        cell,
        status
    ):

        status = str(
            status
        ).strip().lower()

        if status in (
            "pass",
            "başarılı"
        ):

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="C6EFCE"
            )

            cell.font = Font(
                color="006100",
                bold=True
            )

        elif status in (
            "fail",
            "başarısız"
        ):

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FFC7CE"
            )

            cell.font = Font(
                color="9C0006",
                bold=True
            )

        elif status in (
            "blocked",
            "engellendi"
        ):

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FFEB9C"
            )

            cell.font = Font(
                color="9C6500",
                bold=True
            )

        else:

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="E7E6E6"
            )

            cell.font = Font(
                color="666666",
                bold=True
            )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # =================================================
    # COPY STYLE
    # =================================================

    def copy_style(
        self,
        source,
        target
    ):

        if source.has_style:

            target._style = copy(
                source._style
            )

        if source.number_format:

            target.number_format = (
                source.number_format
            )

        if source.alignment:

            target.alignment = copy(
                source.alignment
            )

        if source.font:

            target.font = copy(
                source.font
            )

        if source.fill:

            target.fill = copy(
                source.fill
            )

        if source.border:

            target.border = copy(
                source.border
            )

        if source.protection:

            target.protection = copy(
                source.protection
            )

    # =================================================
    # COPY ROW STYLE
    # =================================================

    def copy_row_style(
        self,
        sheet,
        source_row,
        target_row
    ):

        for col in range(1, 5):

            source = sheet.cell(
                row=source_row,
                column=col
            )

            target = sheet.cell(
                row=target_row,
                column=col
            )

            self.copy_style(
                source,
                target
            )

    # =================================================
    # SECTION HEADER
    # =================================================

    def create_section_header(
        self,
        sheet,
        row,
        text
    ):

        fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        font = Font(
            color="FFFFFF",
            bold=True,
            size=11
        )

        alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        for col in range(1, 5):

            cell = sheet.cell(
                row=row,
                column=col
            )

            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

        sheet.cell(
            row=row,
            column=1
        ).value = text

        sheet.row_dimensions[
            row
        ].height = 22

    # =================================================
    # STEP HEADERS
    # =================================================

    def create_step_headers(
        self,
        sheet,
        row
    ):

        headers = [

            "Adım No",

            "İşlem",

            "Test Verisi",

            "Beklenen Sonuç"

        ]

        fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        font = Font(
            color="000000",
            bold=True,
            size=11
        )

        alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        for col, value in enumerate(
            headers,
            start=1
        ):

            cell = sheet.cell(
                row=row,
                column=col
            )

            cell.value = value
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

        sheet.row_dimensions[
            row
        ].height = 22

    # =================================================
    # LABEL STYLE
    # =================================================

    def apply_label_style(
        self,
        cell
    ):

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        cell.font = Font(
            color="000000",
            bold=True,
            size=11
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

    # =================================================
    # EXPORT
    # =================================================

    def export(self):

        workbook = load_workbook(
            TEMPLATE_PATH
        )

        # =================================================
        # SUMMARY SHEET
        # =================================================

        summary = workbook[
            "Test Senaryoları"
        ]

        # =================================================
        # SUMMARY HEADERS
        # =================================================

        headers = [

            "Test ID",

            "Test Adı",

            "Öncelik",

            "Modül",

            "Test Türü",

            "Durum",

            "Test Ortamı",

            "Oluşturulma Tarihi"

        ]

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
            size=11
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        for col, header in enumerate(
            headers,
            start=1
        ):

            cell = summary.cell(
                row=1,
                column=col
            )

            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        summary.row_dimensions[
            1
        ].height = 30

        # =================================================
        # ESKİ TESTLERİ TEMİZLE
        # =================================================

        if summary.max_row > 1:

            summary.delete_rows(
                2,
                summary.max_row - 1
            )

        # =================================================
        # TESTLER
        # =================================================

        tests = self.db.get_all_test_cases()

        row = 2

        for test in tests:

            # =================================================
            # DATABASE INDEXLERİ
            # =================================================
            #
            # 0  = tc_id
            # 1  = name
            # 2  = priority
            # 3  = application
            # 4  = version
            # 5  = creator
            # 6  = create_date
            # 7  = automation
            # 8  = status
            # 9  = automation_requested
            # 10 = automation_completed
            # 11 = automation_scenario
            # 12 = test_type
            # 13 = test_environment
            # 14 = error_code
            # 15 = error_priority

            tc_id = test[0]

            # =================================================
            # SUMMARY
            # =================================================

            # Test ID

            cell = summary.cell(
                row=row,
                column=1
            )

            cell.value = tc_id

            cell.hyperlink = (
                f"#'{tc_id}'!A1"
            )

            cell.style = "Hyperlink"

            # Test Adı

            summary.cell(
                row=row,
                column=2
            ).value = test[1]

            # Öncelik

            summary.cell(
                row=row,
                column=3
            ).value = test[2]

            # Modül

            summary.cell(
                row=row,
                column=4
            ).value = MODULE_TRANSLATIONS.get(
                test[3],
                test[3]
            )

            # Test Türü

            test_type = (
                test[12]
                if len(test) > 12
                and test[12]
                else "-"
            )

            summary.cell(
                row=row,
                column=5
            ).value = test_type

            # Durum

            status = STATUS_TRANSLATIONS.get(
                test[8],
                test[8]
            )

            status_cell = summary.cell(
                row=row,
                column=6
            )

            status_cell.value = status

            self.apply_status_style(
                status_cell,
                status
            )

            # Test Ortamı

            test_environment = (
                test[13]
                if len(test) > 13
                and test[13]
                else "-"
            )

            summary.cell(
                row=row,
                column=7
            ).value = test_environment

            # Oluşturulma Tarihi

            summary.cell(
                row=row,
                column=8
            ).value = test[6]

            # =================================================
            # HÜCRE HİZALAMA
            # =================================================

            for col in range(1, 9):

                summary.cell(
                    row=row,
                    column=col
                ).alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )

            row += 1

            # =================================================
            # DETAY SAYFASI
            # =================================================

            if tc_id in workbook.sheetnames:

                del workbook[
                    tc_id
                ]

            template = workbook[
                "Template"
            ]

            detail = workbook.copy_worksheet(
                template
            )

            detail.title = tc_id

            # =================================================
            # KOLON GENİŞLİKLERİ
            # =================================================

            detail.column_dimensions[
                "A"
            ].width = 28

            detail.column_dimensions[
                "B"
            ].width = 40

            detail.column_dimensions[
                "C"
            ].width = 40

            detail.column_dimensions[
                "D"
            ].width = 40

            # =================================================
            # DATABASE
            # =================================================

            pre = self.db.get_pre_conditions(
                tc_id
            )

            data = self.db.get_test_data(
                tc_id
            )

            steps = self.db.get_test_steps(
                tc_id
            )

            post = self.db.get_post_conditions(
                tc_id
            )

            # =================================================
            # GERİ DÖN
            # =================================================

            detail["A1"] = (
                "← Test Listesine Dön"
            )

            detail["A1"].hyperlink = (
                "#'Test Senaryoları'!A1"
            )

            detail["A1"].style = (
                "Hyperlink"
            )

            # =================================================
            # TEST BİLGİLERİ
            # =================================================

            labels = [

                "Test ID",

                "Test Adı",

                "Öncelik",

                "Modül",

                "Versiyon",

                "Durum",

                "Oluşturan",

                "Oluşturulma Tarihi",

                "Test Türü",

                "Test Ortamı"

            ]

            values = [

                tc_id,

                test[1],

                test[2],

                MODULE_TRANSLATIONS.get(
                    test[3],
                    test[3]
                ),

                test[4],

                STATUS_TRANSLATIONS.get(
                    test[8],
                    test[8]
                ),

                test[5],

                test[6],

                (
                    test[12]
                    if len(test) > 12
                    and test[12]
                    else "-"
                ),

                (
                    test[13]
                    if len(test) > 13
                    and test[13]
                    else "-"
                )

            ]

            for i, (
                label,
                value
            ) in enumerate(
                zip(labels, values),
                start=2
            ):

                label_cell = detail.cell(
                    row=i,
                    column=1
                )

                value_cell = detail.cell(
                    row=i,
                    column=2
                )

                label_cell.value = label

                value_cell.value = value

                self.apply_label_style(
                    label_cell
                )

                value_cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

                # Durum rengi

                if label == "Durum":

                    self.apply_status_style(
                        value_cell,
                        test[8]
                    )

            # =================================================
            # OTOMASYON BİLGİLERİ
            # =================================================

            current_row = 13

            self.create_section_header(
                detail,
                current_row,
                "Otomasyon Bilgileri"
            )

            current_row += 1

            automation_labels = [

                "Otomasyonlaştırılsın mı?",

                "Otomasyonlaştırıldı mı?",

                "Otomasyon Senaryo Karşılığı"

            ]

            automation_values = [

                (
                    "Evet"
                    if len(test) > 9
                    and test[9]
                    else "Hayır"
                ),

                (
                    "Evet"
                    if len(test) > 10
                    and test[10]
                    else "Hayır"
                ),

                (
                    test[11]
                    if len(test) > 11
                    and test[11]
                    else "-"
                )

            ]

            for label, value in zip(
                automation_labels,
                automation_values
            ):

                detail.cell(
                    row=current_row,
                    column=1
                ).value = label

                detail.cell(
                    row=current_row,
                    column=2
                ).value = value

                self.apply_label_style(
                    detail.cell(
                        row=current_row,
                        column=1
                    )
                )

                detail.cell(
                    row=current_row,
                    column=2
                ).alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

                current_row += 1

            current_row += 1

            # =================================================
            # HATA BİLGİLERİ
            # =================================================

            self.create_section_header(
                detail,
                current_row,
                "Hata Bilgileri"
            )

            current_row += 1

            error_labels = [

                "Hata Kodu",

                "Hata Önceliği"

            ]

            error_values = [

                (
                    test[14]
                    if len(test) > 14
                    and test[14]
                    else "-"
                ),

                (
                    test[15]
                    if len(test) > 15
                    and test[15]
                    else "Yok"
                )

            ]

            for label, value in zip(
                error_labels,
                error_values
            ):

                detail.cell(
                    row=current_row,
                    column=1
                ).value = label

                detail.cell(
                    row=current_row,
                    column=2
                ).value = value

                self.apply_label_style(
                    detail.cell(
                        row=current_row,
                        column=1
                    )
                )

                detail.cell(
                    row=current_row,
                    column=2
                ).alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

                current_row += 1

            current_row += 2

            # =================================================
            # PRE CONDITIONS
            # =================================================

            self.create_section_header(
                detail,
                current_row,
                "Ön Koşullar"
            )

            current_row += 1

            if not pre:

                detail.cell(
                    row=current_row,
                    column=2
                ).value = (
                    "Ön koşul bulunmuyor."
                )

                current_row += 1

            else:

                for item in pre:

                    detail.cell(
                        row=current_row,
                        column=2
                    ).value = item[0]

                    detail.cell(
                        row=current_row,
                        column=2
                    ).alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                    current_row += 1

            current_row += 1

            # =================================================
            # TEST DATA
            # =================================================

            self.create_section_header(
                detail,
                current_row,
                "Test Verileri"
            )

            current_row += 1

            if not data:

                detail.cell(
                    row=current_row,
                    column=2
                ).value = (
                    "Test verisi bulunmuyor."
                )

                current_row += 1

            else:

                for item in data:

                    detail.cell(
                        row=current_row,
                        column=2
                    ).value = item[0]

                    detail.cell(
                        row=current_row,
                        column=3
                    ).value = item[1]

                    detail.cell(
                        row=current_row,
                        column=2
                    ).alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                    detail.cell(
                        row=current_row,
                        column=3
                    ).alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                    current_row += 1

            current_row += 1

            # =================================================
            # TEST STEPS
            # =================================================

            self.create_section_header(
                detail,
                current_row,
                "Test Adımları"
            )

            current_row += 1

            self.create_step_headers(
                detail,
                current_row
            )

            current_row += 1

            if not steps:

                detail.cell(
                    row=current_row,
                    column=1
                ).value = "-"

                detail.cell(
                    row=current_row,
                    column=2
                ).value = (
                    "Test adımı bulunmuyor."
                )

                current_row += 1

            else:

                for step in steps:

                    detail.cell(
                        row=current_row,
                        column=1
                    ).value = step[0]

                    detail.cell(
                        row=current_row,
                        column=2
                    ).value = step[1]

                    detail.cell(
                        row=current_row,
                        column=3
                    ).value = step[2]

                    detail.cell(
                        row=current_row,
                        column=4
                    ).value = step[3]

                    for col in range(1, 5):

                        detail.cell(
                            row=current_row,
                            column=col
                        ).alignment = Alignment(
                            vertical="top",
                            wrap_text=True
                        )

                    current_row += 1

            current_row += 1

            # =================================================
            # POST CONDITIONS
            # =================================================

            self.create_section_header(
                detail,
                current_row,
                "Son Koşullar"
            )

            current_row += 1

            if not post:

                detail.cell(
                    row=current_row,
                    column=2
                ).value = (
                    "Son koşul bulunmuyor."
                )

                current_row += 1

            else:

                for item in post:

                    detail.cell(
                        row=current_row,
                        column=2
                    ).value = item[0]

                    detail.cell(
                        row=current_row,
                        column=2
                    ).alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                    current_row += 1

        # =================================================
        # SUMMARY COLUMN WIDTHS
        # =================================================

        summary.column_dimensions[
            "A"
        ].width = 18

        summary.column_dimensions[
            "B"
        ].width = 40

        summary.column_dimensions[
            "C"
        ].width = 15

        summary.column_dimensions[
            "D"
        ].width = 20

        summary.column_dimensions[
            "E"
        ].width = 20

        summary.column_dimensions[
            "F"
        ].width = 18

        summary.column_dimensions[
            "G"
        ].width = 40

        summary.column_dimensions[
            "H"
        ].width = 20

        # =================================================
        # TEMPLATE'İ EN SONA TAŞI
        # =================================================

        template = workbook[
            "Template"
        ]

        workbook._sheets.remove(
            template
        )

        workbook._sheets.append(
            template
        )

        template.sheet_state = "hidden"

        # =================================================
        # KAYDET
        # =================================================

        workbook.save(
            EXPORT_PATH
        )

        # =================================================
        # DATABASE KAPAT
        # =================================================

        if self.owns_db:

            self.db.close()